from anyio import key
import cv2
import numpy as np
import psutil
import time
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import cProfile
import pstats
from concurrent.futures import ThreadPoolExecutor
from collections import Counter
import json
import xml.etree.ElementTree as ET
import matplotlib.pyplot as plt
import seaborn as sns


with open("user_input_data.json", "r") as file:
    config = json.load(file)

video_path = config["video"]
color_channel = config["color_channel"]
rows, cols = config["grids"]["rows"], config["grids"]["cols"]
pixels_per_meter = config.get("pixels_per_meter", 0)  # 0 => speed estimation disabled
xml_path = config.get("xml_path", "")

roi1_x, roi1_y, roi1_width, roi1_height = 47, 202, 441, 141
roi2_x, roi2_y, roi2_width, roi2_height = 500, 202, 441, 141

num_rows, num_cols = rows, cols

frame_count = 0
start_time = time.time()

cap = cv2.VideoCapture(video_path)

# Read actual FPS from the video; fall back to 30 if unavailable
_raw_fps = cap.get(cv2.CAP_PROP_FPS)
FPS = _raw_fps if _raw_fps and _raw_fps > 0 else 30

# Read the first two frames BEFORE creating the writer, so the writer can be
# sized from the real decoded frame instead of a possibly-wrong cap.get().
ret1, frame1 = cap.read()
ret, frame2 = cap.read()

if not ret1 or frame1 is None:
    raise RuntimeError(f"Could not read any frames from video: {video_path}")

frame_height, frame_width = frame1.shape[:2]

excel_file_path = 'result_matrix1.xlsx'

fourcc = cv2.VideoWriter_fourcc(*'avc1')
out = cv2.VideoWriter('test_output1.mp4', fourcc, FPS, (frame_width, frame_height))
if not out.isOpened():
    print("[WARN] avc1 codec failed, falling back to mp4v")
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter('test_output1.mp4', fourcc, FPS, (frame_width, frame_height))

OUTPUT_WRITER_OK = out.isOpened()
if not OUTPUT_WRITER_OK:
    print(f"[ERROR] VideoWriter failed to open for 'test_output1.mp4' "
          f"(fps={FPS}, size=({frame_width}, {frame_height})). "
          "Processed frames will be shown (if a display is available) but NOT saved to disk.")

# Whether a GUI display is available; probed lazily on the first imshow call
# in main(), so this also correctly detects headless environments.
DISPLAY_ENABLED = True

# ─────────────────────────────────────────────────────────────────────────────
# Persistent, reusable thread pools.
#
# The original code created a brand-new ThreadPoolExecutor() on every single
# frame (inside apply_histogram_equalization_parallel and LaneTracker.process_
# grid), and nested a lane-level pool around a cell-level pool. That means
# thread creation/teardown overhead every frame, plus two pools competing for
# the same cores. Two long-lived pools fix both problems:
#
#   _CELL_POOL  - fine-grained, GIL-releasing OpenCV work (per grid cell,
#                 per-ROI histogram equalization). Sized larger since this is
#                 where the real native/parallel work happens.
#   _LANE_POOL  - coarse-grained, lane-level dispatch (2 lanes at a time).
#                 Kept as a SEPARATE pool so lane-level tasks that internally
#                 submit to _CELL_POOL never wait on a worker slot in the same
#                 pool they're blocked on (avoids self-nesting deadlock risk).
# ─────────────────────────────────────────────────────────────────────────────
_CELL_POOL = ThreadPoolExecutor(max_workers=min(32, (os.cpu_count() or 4) * 4))
_LANE_POOL = ThreadPoolExecutor(max_workers=2)


# ─────────────────────────────────────────────────────────────────────────────
# Shared, stateless helpers
# ─────────────────────────────────────────────────────────────────────────────
def process_grid_channel(channel, grid_width, grid_height):
    """Contour extraction for a single grid cell (per-cell motion detection)."""
    blur = cv2.GaussianBlur(channel, (5, 5), 1)
    _, thresh = cv2.threshold(blur, 60, 255, cv2.THRESH_BINARY)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(1, grid_width), max(1, grid_height // 2)))
    dilated = cv2.dilate(thresh, kernel, iterations=5)
    contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    return contours


def apply_histogram_equalization_single(frame, roi_x, roi_y, roi_width, roi_height):
    roi = frame[roi_y:roi_y + roi_height, roi_x:roi_x + roi_width]
    roi_equalized = cv2.equalizeHist(cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY))
    frame[roi_y:roi_y + roi_height, roi_x:roi_x + roi_width] = cv2.cvtColor(roi_equalized, cv2.COLOR_GRAY2BGR)
    return frame


def apply_histogram_equalization_parallel(frame):
    futures = [
        _CELL_POOL.submit(apply_histogram_equalization_single, frame, roi1_x, roi1_y, roi1_width, roi1_height),
        _CELL_POOL.submit(apply_histogram_equalization_single, frame, roi2_x, roi2_y, roi2_width, roi2_height)
    ]
    for future in futures:
        frame = future.result()
    return frame


def process_hsv(frame1, frame2, channels):
    frame1 = apply_histogram_equalization_parallel(frame1)
    frame2 = apply_histogram_equalization_parallel(frame2)

    diff = cv2.absdiff(frame1, frame2)
    hsv = cv2.cvtColor(diff, cv2.COLOR_BGR2HSV)
    channels_data = [cv2.split(hsv)[i] for i in channels]
    return channels_data


def process_grayscale(frame1, frame2):
    frame1 = apply_histogram_equalization_parallel(frame1)
    frame2 = apply_histogram_equalization_parallel(frame2)

    diff = cv2.absdiff(frame1, frame2)
    gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
    return [gray]


user_choice = color_channel
choices = {
    'H': [0], 'S': [1], 'V': [2],
    'H+S': [0, 1], 'H+V': [0, 2], 'S+V': [1, 2], 'H+S+V': [0, 1, 2],
    'gray': 'gray'
}
channels = choices[user_choice]


# ─────────────────────────────────────────────────────────────────────────────
# DETRAC ground-truth evaluation helpers (parameterized by ROI/grid so the
# same functions serve either lane).
# ─────────────────────────────────────────────────────────────────────────────
def map_detrac_label(vehicle_type):
    mapping = {
        'car': 'Car',
        'van': 'Van',
        'bus': 'Bus',
        'others': 'Unknown'
    }
    return mapping.get(vehicle_type.lower(), 'Unknown')


def load_detrac(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    gt_data = {}

    for frame in root.findall('frame'):
        frame_id = int(frame.get('num'))
        gt_data[frame_id] = []

        target_list = frame.find('target_list')
        if target_list is None:
            continue

        for target in target_list.findall('target'):
            box = target.find('box')
            attr = target.find('attribute')

            if box is None or attr is None:
                continue

            x = float(box.get('left'))
            y = float(box.get('top'))
            w = float(box.get('width'))
            h = float(box.get('height'))

            label = map_detrac_label(attr.get('vehicle_type', 'others'))
            obj_id = int(target.get('id', -1))

            gt_data[frame_id].append({
                "id": obj_id,
                "bbox": (x, y, w, h),
                "label": label
            })

    return gt_data


def filter_gt_boxes_to_roi(gt_boxes, roi_x, roi_y, roi_width, roi_height):
    return [
        obj for obj in gt_boxes
        if not (
            obj["bbox"][0] + obj["bbox"][2] < roi_x or
            obj["bbox"][0] > roi_x + roi_width or
            obj["bbox"][1] + obj["bbox"][3] < roi_y or
            obj["bbox"][1] > roi_y + roi_height
        )
    ]


def create_gt_label_grid(gt_boxes, roi_x, roi_y, grid_width, grid_height, num_rows, num_cols):
    grid = np.full((num_rows, num_cols), "Empty", dtype=object)

    for obj in gt_boxes:
        x, y, w, h = obj["bbox"]
        label = obj["label"]

        # ignore unknown classes
        if label not in ["Car", "Van", "Bus"]:
            continue

        cx = x + w / 2
        cy = y + h / 2

        col = int((cx - roi_x) / grid_width)
        row = int((cy - roi_y) / grid_height)

        if 0 <= row < num_rows and 0 <= col < num_cols:
            grid[row][col] = label

    return grid


def match_vehicle(pred_bbox, gt_boxes):
    best_id = None
    max_iou = 0

    px, py, pw, ph = pred_bbox

    for obj in gt_boxes:
        gx, gy, gw, gh = obj["bbox"]

        ix = max(0, min(px + pw, gx + gw) - max(px, gx))
        iy = max(0, min(py + ph, gy + gh) - max(py, gy))
        inter = ix * iy

        union = pw * ph + gw * gh - inter
        iou = inter / union if union > 0 else 0

        if iou > max_iou:
            max_iou = iou
            best_id = obj.get("id")

    return best_id if max_iou > 0.1 else None


def compute_vehicle_metrics(cm):
    metrics = {}
    classes = ["Car", "Van", "Bus"]

    for i, cls in enumerate(classes):
        tp = cm[i][i]
        fp = cm[:, i].sum() - tp
        fn = cm[i, :].sum() - tp

        precision = tp / (tp + fp + 1e-6)
        recall = tp / (tp + fn + 1e-6)
        f1 = 2 * precision * recall / (precision + recall + 1e-6)

        metrics[cls] = (precision, recall, f1)

    return metrics


# ─────────────────────────────────────────────────────────────────────────────
# LaneTracker — encapsulates everything seq.py did globally, per ROI/lane.
# Two independent instances (lane1, lane2) let both lanes keep their own
# vehicle tracks, counts, densities, and speed logs while grid detection
# for each lane still runs across a thread pool internally, and the two
# lanes' grid detection is itself run in parallel.
# ─────────────────────────────────────────────────────────────────────────────
class LaneTracker:
    def __init__(self, name, roi_x, roi_y, roi_width, roi_height, num_rows, num_cols,
                 fps, pixels_per_meter, hud_x_offset, hud_y_offset, box_color):
        self.name = name
        self.roi_x = roi_x
        self.roi_y = roi_y
        self.roi_width = roi_width
        self.roi_height = roi_height
        self.num_rows = num_rows
        self.num_cols = num_cols
        self.grid_width = roi_width // num_cols
        self.grid_height = roi_height // num_rows
        self.fps = fps
        self.pixels_per_meter = pixels_per_meter
        self.hud_x_offset = hud_x_offset
        self.hud_y_offset = hud_y_offset
        self.box_color = box_color

        # Tracking state
        self.vehicles = {}
        self.vehicle_history = []      # archived (finished) tracks
        self.next_vehicle_id = 0
        self.vehicle_count = 0
        self.vehicle_count_history = []
        self.recent_crossings = []
        self.speed_log = []

        # Per-frame snapshots retained for DETRAC ground-truth evaluation and
        # Excel export — mirrors seq.py's all_frames_data / all_frames_vehicle_data.
        self.all_frames_data = []            # (frame_count, binary_matrix, label_matrix)
        self.all_frames_vehicle_data = []    # (frame_count, {vid: {"bbox":..., "label":...}})

        # Density state
        self.density_values = []
        self.combined_density_values = []
        self.base_low = 0.25
        self.base_high = 0.55

        # Tunables
        self.count_line_y = roi_y + roi_height // 2
        self.min_speed_dist_px = 3 * self.grid_height
        self.cross_cooldown_frames = 30
        self.cross_cooldown_px = 60
        self.label_ewa_alpha = 0.15
        self.min_frames_for_label = 5

    # ── grid / motion detection ─────────────────────────────────────────────
    def process_grid_cell(self, args):
        row, col, frame1, channels_data = args
        grid_x = self.roi_x + col * self.grid_width
        grid_y = self.roi_y + row * self.grid_height
        grid_frame = frame1[grid_y:grid_y + self.grid_height, grid_x:grid_x + self.grid_width]
        if grid_frame.size == 0:
            return row, col, 0, []

        detection_flags = []
        cell_contours = []
        for channel in channels_data:
            grid_channel = channel[grid_y:grid_y + self.grid_height, grid_x:grid_x + self.grid_width]
            contours = process_grid_channel(grid_channel, self.grid_width, self.grid_height)
            valid_contours = [c for c in contours if cv2.contourArea(c) > 100]
            detection_flags.append(len(valid_contours) > 0)
            cell_contours.extend(valid_contours)

        hit = 1 if detection_flags and all(detection_flags) else 0
        return row, col, hit, cell_contours

    def process_grid(self, frame1, channels_data):
        result_matrix = np.zeros((self.num_rows, self.num_cols), dtype=int)
        combined_contours = []
        args_list = [(row, col, frame1, channels_data)
                     for row in range(self.num_rows) for col in range(self.num_cols)]
        for row, col, hit, cell_contours in _CELL_POOL.map(self.process_grid_cell, args_list):
            if hit:
                result_matrix[row, col] = 1
                combined_contours.append((row, col, cell_contours))
        return result_matrix, combined_contours

    def draw_grid_overlay(self, frame1, result_matrix):
        for row in range(self.num_rows):
            for col in range(self.num_cols):
                grid_x = self.roi_x + col * self.grid_width
                grid_y = self.roi_y + row * self.grid_height
                color = (0, 255, 0) if result_matrix[row, col] == 1 else (0, 0, 255)
                cv2.rectangle(frame1, (grid_x, grid_y),
                               (grid_x + self.grid_width, grid_y + self.grid_height), color, 2)

    # ── clustering hot cells into vehicle-sized objects ─────────────────────
    def group_cells(self, cell_contours):
        """4-connected flood-fill grouping of hot grid cells."""
        grid = np.zeros((self.num_rows, self.num_cols), dtype=int)
        for (r, c, _) in cell_contours:
            grid[r][c] = 1

        visited = np.zeros_like(grid)
        clusters = []

        for r in range(self.num_rows):
            for c in range(self.num_cols):
                if grid[r][c] == 1 and visited[r][c] == 0:
                    stack = [(r, c)]
                    cluster = []
                    while stack:
                        rr, cc = stack.pop()
                        if rr < 0 or rr >= self.num_rows or cc < 0 or cc >= self.num_cols:
                            continue
                        if visited[rr][cc] or grid[rr][cc] == 0:
                            continue
                        visited[rr][cc] = 1
                        cluster.append((rr, cc))
                        stack.extend([(rr + 1, cc), (rr - 1, cc), (rr, cc + 1), (rr, cc - 1)])
                    clusters.append(cluster)

        return clusters

    def cluster_to_object(self, cluster):
        cols = [c for (_, c) in cluster]
        rows = [r for (r, _) in cluster]
        col_span = len(set(cols))
        row_span = len(set(rows))

        xs, ys = [], []
        for r, c in cluster:
            gx = self.roi_x + c * self.grid_width
            gy = self.roi_y + r * self.grid_height
            xs.extend([gx, gx + self.grid_width])
            ys.extend([gy, gy + self.grid_height])

        bx, by = min(xs), min(ys)
        bw = max(xs) - bx
        bh = max(ys) - by
        cx = bx + bw // 2
        cy = by + bh // 2

        return {
            "centroid": (cx, cy),
            "bbox": (bx, by, bw, bh),
            "col_span": col_span,
            "row_span": row_span,
            "cell_count": len(cluster),
        }

    def classify_by_grid(self, col_span, row_span):
        norm_col = col_span / self.num_cols
        norm_row = row_span / self.num_rows

        # Bus: wide OR very large overall
        if norm_col > 0.75 or (norm_col > 0.55 and norm_row > 0.55):
            return "Bus"
        # Van: long but not too wide
        if norm_row > 0.45:
            return "Van"
        return "Car"

    # ── crossing-line de-duplication ────────────────────────────────────────
    def is_duplicate_crossing(self, cx, frame_count):
        for past_frame, past_cx in self.recent_crossings:
            if (frame_count - past_frame) < self.cross_cooldown_frames and abs(cx - past_cx) < self.cross_cooldown_px:
                return True
        return False

    def prune_crossings(self, frame_count):
        self.recent_crossings[:] = [
            (f, cx) for (f, cx) in self.recent_crossings
            if (frame_count - f) < self.cross_cooldown_frames
        ]

    def suppress_duplicate_tracks(self, frame_count):
        """Suppress duplicate active tracks that overlap heavily in the same frame."""
        active = {
            k: v for k, v in self.vehicles.items()
            if v.get('last_seen') == frame_count and v.get('bbox') is not None
        }
        keys = list(active.keys())
        suppressed = set()

        for i in range(len(keys)):
            for j in range(i + 1, len(keys)):
                ka, kb = keys[i], keys[j]
                if ka in suppressed or kb in suppressed:
                    continue

                ax, ay, aw, ah = active[ka]['bbox']
                bx, by, bw, bh = active[kb]['bbox']

                ix = max(0, min(ax + aw, bx + bw) - max(ax, bx))
                iy = max(0, min(ay + ah, by + bh) - max(ay, by))
                inter = ix * iy
                union = aw * ah + bw * bh - inter
                iou = inter / union if union > 0 else 0

                if iou > 0.40:
                    age_a = active[ka].get('age', 0)
                    age_b = active[kb].get('age', 0)
                    if age_a > age_b:
                        loser = kb
                    elif age_b > age_a:
                        loser = ka
                    else:
                        loser = max(ka, kb)

                    suppressed.add(loser)
                    self.vehicles[loser]['counted'] = True
                    self.vehicles[loser]['last_seen'] = -1

        return suppressed

    def _archive_and_remove(self, key):
        v = self.vehicles.get(key)
        if v is not None:
            self.vehicle_history.append({
                "vehicle_id": key,
                "label": v.get("final_label", "Unknown"),
                "avg_speed_kmph": round(float(np.mean(v["speed_history"])), 1) if v.get("speed_history") else 0.0,
                "counted": v.get("counted", False),
            })
            del self.vehicles[key]

    # ── main tracking / classification / speed / counting step ─────────────
    #
    # This is split into compute_tracking() (pure state update -- no frame
    # writes) and draw_tracking_results() (all cv2 drawing). The two lanes'
    # compute_tracking() calls only touch their own LaneTracker instance
    # attributes (self.vehicles, self.vehicle_count, self.speed_log, ...), so
    # they can safely run concurrently in separate threads. draw_tracking_
    # results() writes onto the shared frame1 buffer, so it must be called
    # for each lane AFTER both lanes' compute_tracking() calls have finished
    # (see main()), never concurrently with another lane's draw call.
    #
    # Caveat: compute_tracking() is mostly pure-Python dict/list bookkeeping,
    # not GIL-releasing OpenCV/NumPy work on large arrays, so running lane1's
    # and lane2's compute_tracking() in separate threads does not guarantee
    # true simultaneous CPU execution the way threading the grid-detection
    # (OpenCV) stage does. The benefit here is architectural: safe, race-free
    # separation of "update state" from "draw state", not a raw speed boost.
    def compute_tracking(self, frame1_shape, frame_count, cell_contours):
        """frame1_shape: the (height, width, channels) shape of the current
        frame, used only to normalize a vertical-position weight below. The
        pixel buffer itself is never touched here -- see class docstring
        above for why that matters."""
        keys_to_remove = []
        MATCH_THRESHOLD = self.roi_height // 2
        draw_boxes = []

        # HUD text is deferred to draw_tracking_results(); just record the
        # count that should be displayed for this frame.
        self._pending_draw = {"hud_count": self.vehicle_count, "boxes": []}

        self.prune_crossings(frame_count)

        if not cell_contours:
            self.vehicle_count_history.append(0)
            for key in list(self.vehicles.keys()):
                if frame_count - self.vehicles[key]['last_seen'] > self.fps * 2:
                    keys_to_remove.append(key)
            for key in keys_to_remove:
                self._archive_and_remove(key)
            return

        # STEP 1: group hot cells into spatially-connected clusters
        clusters = self.group_cells(cell_contours)

        # STEP 2: convert each cluster -> object dict
        objects = [self.cluster_to_object(cl) for cl in clusters]

        # STEP 3: match each object to an existing track (nearest centroid)
        matched_track_ids = set()

        for obj in objects:
            cx, cy = obj["centroid"]
            best_match = None
            min_dist = MATCH_THRESHOLD

            for vid, v in self.vehicles.items():
                if vid in matched_track_ids:
                    continue
                px, py = v["centroid"]
                d = np.sqrt((cx - px) ** 2 + (cy - py) ** 2)
                if d < min_dist:
                    min_dist = d
                    best_match = vid

            if best_match is not None:
                v = self.vehicles[best_match]
                old_centroid = v["centroid"]
                new_centroid = obj["centroid"]

                # speed estimation (accumulate until displacement >= threshold)
                new_age = v.get("age", 0) + 1
                if new_age > 1:
                    last_sc = v.get("last_speed_centroid", new_centroid)
                    if new_centroid != last_sc:
                        dx = new_centroid[0] - last_sc[0]
                        dy = new_centroid[1] - last_sc[1]
                        dist_px = np.hypot(dx, dy)
                        if dist_px >= self.min_speed_dist_px:
                            dt = (frame_count - v.get("last_speed_frame", frame_count)) / self.fps
                            if dt > 0 and self.pixels_per_meter > 0:
                                speed_kmph = (dist_px / self.pixels_per_meter) / dt * 3.6
                                hist = v.setdefault("speed_history", [])
                                hist.append(speed_kmph)
                                if len(hist) > 15:
                                    hist.pop(0)
                                v["speed_kmph"] = round(float(np.mean(hist)), 1)
                                v["speed_initialized"] = True
                            v["last_speed_centroid"] = new_centroid
                            v["last_speed_frame"] = frame_count

                v["prev_centroid"] = old_centroid
                v["centroid"] = new_centroid
                v["bbox"] = obj["bbox"]
                v["col_span"] = obj["col_span"]
                v["row_span"] = obj["row_span"]
                v["cell_count"] = obj["cell_count"]
                v["last_seen"] = frame_count
                v["age"] = new_age

                matched_track_ids.add(best_match)
            else:
                self.vehicles[self.next_vehicle_id] = {
                    "centroid": obj["centroid"],
                    "prev_centroid": obj["centroid"],
                    "bbox": obj["bbox"],
                    "col_span": obj["col_span"],
                    "row_span": obj["row_span"],
                    "cell_count": obj["cell_count"],
                    "last_seen": frame_count,
                    "age": 1,
                    "counted": False,
                    "frame_age": 0,
                    "recent_labels": [],
                    "weighted_votes": {"Car": 0.0, "Van": 0.0, "Bus": 0.0},
                    "final_label": "Unknown",
                    "display_label": "Unknown",
                    "last_speed_centroid": obj["centroid"],
                    "last_speed_frame": frame_count,
                    "speed_history": [],
                    "speed_kmph": 0.0,
                    "speed_initialized": False,
                }
                matched_track_ids.add(self.next_vehicle_id)
                self.next_vehicle_id += 1

        # STEP 4: classify each active track + log speed data
        for key, v in self.vehicles.items():
            if v["last_seen"] != frame_count:
                continue

            v["frame_age"] = v.get("frame_age", 0) + 1
            raw_label = self.classify_by_grid(v["col_span"], v["row_span"])

            wv = v.setdefault("weighted_votes", {"Car": 0.0, "Van": 0.0, "Bus": 0.0})
            cy = v["centroid"][1]
            pos_weight = cy / max(1, frame1_shape[0])
            time_weight = (1 + self.label_ewa_alpha) ** v["frame_age"]
            weight = pos_weight * time_weight
            wv[raw_label] = wv.get(raw_label, 0.0) + weight
            v["final_label"] = max(wv, key=wv.get)

            labels = v.setdefault("recent_labels", [])
            labels.append(raw_label)
            if len(labels) > 10:
                labels.pop(0)
            v["display_label"] = max(set(labels), key=labels.count)

            if v.get("bbox") is not None:
                bx, by, bw, bh = v["bbox"]
                status = v["display_label"]
                label_y = max(by - 5 - (key % 3) * 14, 12)
                speed_str = f" {v['speed_kmph']}km/h" if v.get("speed_kmph", 0) > 0 else ""
                draw_boxes.append({
                    "bbox": (bx, by, bw, bh),
                    "text": f"{self.name[:1]}{key} {status}{speed_str}",
                    "label_pos": (bx, label_y),
                })

            if v.get("frame_age", 0) >= self.min_frames_for_label:
                self.speed_log.append({
                    "lane": self.name,
                    "frame": frame_count,
                    "vehicle_id": key,
                    "label": v["final_label"],
                    "speed_kmph": v.get("speed_kmph", 0.0),
                    "centroid_x": v["centroid"][0],
                    "centroid_y": v["centroid"][1],
                })

        self.suppress_duplicate_tracks(frame_count)

        # STEP 5: count vehicles crossing the line
        temp_count = 0
        for key in list(self.vehicles.keys()):
            value = self.vehicles[key]
            if (not value["counted"]
                    and value["age"] >= 2
                    and value["prev_centroid"][1] < self.count_line_y
                    and value["centroid"][1] >= self.count_line_y):

                cx = value["centroid"][0]
                if self.is_duplicate_crossing(cx, frame_count):
                    self.vehicles[key]["counted"] = True
                    continue

                self.vehicle_count += 1
                temp_count += 1
                self.vehicles[key]["counted"] = True
                self.recent_crossings.append((frame_count, cx))
                self.vehicles[key]["last_seen"] = frame_count

        self.vehicle_count_history.append(temp_count)

        # cleanup stale tracks
        for key in list(self.vehicles.keys()):
            if frame_count - self.vehicles[key]["last_seen"] > self.fps * 2:
                keys_to_remove.append(key)
        for key in keys_to_remove:
            self._archive_and_remove(key)

        self._pending_draw = {"hud_count": self.vehicle_count, "boxes": draw_boxes}

    def draw_tracking_results(self, frame1):
        """Draw this lane's HUD text and vehicle boxes/labels onto frame1.
        Must be called only after compute_tracking() has finished for BOTH
        lanes for this frame, since it's the only part of the tracking step
        that touches the shared frame buffer."""
        pending = getattr(self, "_pending_draw", None)
        if pending is None:
            return

        cv2.putText(frame1, f"{self.name} Vehicles Passed: {pending['hud_count']}",
                    (self.hud_x_offset, self.hud_y_offset), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)

        for item in pending["boxes"]:
            bx, by, bw, bh = item["bbox"]
            cv2.rectangle(frame1, (bx, by), (bx + bw, by + bh), self.box_color, 2)
            cv2.putText(frame1, item["text"], item["label_pos"],
                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, self.box_color, 1)

    # ── per-frame grid label matrix + evaluation snapshots ─────────────────
    def build_label_matrix(self, result_matrix):
        """For each occupied grid cell, assign the label of whichever active
        vehicle track overlaps it the most (mirrors seq.py's label_matrix)."""
        label_matrix = np.full((self.num_rows, self.num_cols), "Empty", dtype=object)

        for r in range(self.num_rows):
            for c in range(self.num_cols):
                if result_matrix[r, c] == 0:
                    continue

                gx = self.roi_x + c * self.grid_width
                gy = self.roi_y + r * self.grid_height

                best_label = "Empty"
                max_overlap = 0

                for v in self.vehicles.values():
                    if v.get('bbox') is None:
                        continue

                    bx, by, bw, bh = v['bbox']
                    label = v.get('final_label', "Unknown")

                    overlap_x = max(0, min(bx + bw, gx + self.grid_width) - max(bx, gx))
                    overlap_y = max(0, min(by + bh, gy + self.grid_height) - max(by, gy))
                    overlap_area = overlap_x * overlap_y

                    if overlap_area > max_overlap:
                        max_overlap = overlap_area
                        best_label = label

                if max_overlap > 0:
                    label_matrix[r][c] = best_label

        return label_matrix

    def record_frame_data(self, frame_count, result_matrix):
        """Snapshot this frame's binary/label grids and vehicle predictions,
        for later DETRAC evaluation and Excel export."""
        label_matrix = self.build_label_matrix(result_matrix)
        self.all_frames_data.append((frame_count, result_matrix.copy(), label_matrix.copy()))

        frame_vehicle_preds = {}
        for vid, v in self.vehicles.items():
            if v.get('last_seen') != frame_count or v.get('bbox') is None:
                continue
            pred_label = v.get('final_label') if v.get('frame_age', 0) >= self.min_frames_for_label else None
            if pred_label not in ("Car", "Van", "Bus"):
                pred_label = None
            frame_vehicle_preds[vid] = {
                "bbox": v['bbox'],
                "label": pred_label,
            }
        self.all_frames_vehicle_data.append((frame_count, frame_vehicle_preds))

    # ── adaptive density ─────────────────────────────────────────────────────
    def density_calculation(self, frame1, result_matrix):
        density = np.sum(result_matrix == 1) / (self.num_rows * self.num_cols)
        self.density_values.append(density)

        FLOW_WINDOW = max(1, int(round(2 * self.fps)))
        vehicles_last_1s = sum(self.vehicle_count_history[-FLOW_WINDOW:])
        WINDOW = max(1, int(round(self.fps)))
        smoothed_density = np.mean(self.density_values[-WINDOW:])
        MAX_FLOW = 8  # max vehicles/sec expected for this ROI
        flow_score = min(vehicles_last_1s / MAX_FLOW, 1.0)

        combined_density = 0.7 * smoothed_density + 0.3 * flow_score
        self.combined_density_values.append(combined_density)

        if len(self.combined_density_values) > 50:
            p33 = np.percentile(self.combined_density_values, 33)
            p66 = np.percentile(self.combined_density_values, 66)

            delta_low = np.clip(p33 - self.base_low, -0.1, 0.1)
            delta_high = np.clip(p66 - self.base_high, -0.1, 0.1)

            low_th = self.base_low + delta_low
            high_th = self.base_high + delta_high
            if low_th >= high_th:
                high_th = low_th + 0.05
        else:
            low_th, high_th = 0.3, 0.6

        if combined_density > high_th:
            state = "High"
        elif combined_density > low_th:
            state = "Medium"
        else:
            state = "Low"

        cv2.putText(frame1, f"{self.name} Density: {combined_density:.2f} ({state})",
                    (self.hud_x_offset, self.hud_y_offset + 40), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
        return combined_density, state


# ─────────────────────────────────────────────────────────────────────────────
# Instantiate one tracker per lane, keeping their ROIs/state independent.
# ─────────────────────────────────────────────────────────────────────────────
lane1 = LaneTracker("Lane1", roi1_x, roi1_y, roi1_width, roi1_height, num_rows, num_cols,
                     FPS, pixels_per_meter, hud_x_offset=10, hud_y_offset=100, box_color=(0, 255, 255))
lane2 = LaneTracker("Lane2", roi2_x, roi2_y, roi2_width, roi2_height, num_rows, num_cols,
                     FPS, pixels_per_meter, hud_x_offset=620, hud_y_offset=100, box_color=(255, 0, 255))


def main():
    global frame1, frame2, ret, frame_count, DISPLAY_ENABLED
    WINDOW_NAME = "Parallel Vehicle Detection"

    try:
        while cap.isOpened():
            if not ret:
                break
            frame_count += 1

            if frame1.shape[:2] == frame2.shape[:2]:
                if channels == 'gray':
                    channels_data = process_grayscale(frame1, frame2)
                else:
                    channels_data = process_hsv(frame1, frame2, channels)

                # Grid/motion detection for both lanes in parallel (read-only work,
                # safe to run concurrently since it doesn't draw on frame1 yet).
                # Dispatched via the persistent _LANE_POOL rather than creating a
                # new pool every frame; process_grid() itself uses the separate
                # _CELL_POOL internally, so there's no self-nesting deadlock risk.
                grid_future1 = _LANE_POOL.submit(lane1.process_grid, frame1, channels_data)
                grid_future2 = _LANE_POOL.submit(lane2.process_grid, frame1, channels_data)
                result_matrix1, cell_contours1 = grid_future1.result()
                result_matrix2, cell_contours2 = grid_future2.result()

                # Tracking / classification / speed / counting: compute state
                # for both lanes concurrently (each lane only touches its own
                # LaneTracker attributes, so this is race-free), then draw both
                # lanes' results onto the shared frame1 buffer only after BOTH
                # computations have finished. See LaneTracker.compute_tracking
                # docstring for why this buys safety/clarity rather than a
                # guaranteed CPU speedup.
                track_future1 = _LANE_POOL.submit(lane1.compute_tracking, frame1.shape, frame_count, cell_contours1)
                track_future2 = _LANE_POOL.submit(lane2.compute_tracking, frame1.shape, frame_count, cell_contours2)
                track_future1.result()
                track_future2.result()

                lane1.draw_tracking_results(frame1)
                lane2.draw_tracking_results(frame1)

                # Snapshot per-frame grid/label matrices + vehicle predictions
                # (used later for DETRAC evaluation and Excel export).
                lane1.record_frame_data(frame_count, result_matrix1)
                lane2.record_frame_data(frame_count, result_matrix2)

                # Grid overlays
                lane1.draw_grid_overlay(frame1, result_matrix1)
                lane2.draw_grid_overlay(frame1, result_matrix2)

                # Density + count line
                lane1.density_calculation(frame1, result_matrix1)
                lane2.density_calculation(frame1, result_matrix2)

                cv2.line(frame1, (lane1.roi_x, lane1.count_line_y),
                         (lane1.roi_x + lane1.roi_width, lane1.count_line_y), (255, 255, 255), 2)
                cv2.line(frame1, (lane2.roi_x, lane2.count_line_y),
                         (lane2.roi_x + lane2.roi_width, lane2.count_line_y), (255, 255, 255), 2)

                if frame_count % 100 == 0:  # reduce console spam
                    print(f"Frame {frame_count} | Lane1 passed: {lane1.vehicle_count} "
                          f"| Lane2 passed: {lane2.vehicle_count}")

                cv2.putText(frame1, "Frame: {}".format(frame_count), (10, 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)

                # All drawing/annotation on frame1 is complete at this point —
                # write it to disk and, if a display is available, preview it,
                # before advancing to the next frame.
                if OUTPUT_WRITER_OK:
                    out.write(frame1)

                if DISPLAY_ENABLED:
                    try:
                        cv2.imshow(WINDOW_NAME, frame1)
                    except cv2.error as e:
                        DISPLAY_ENABLED = False
                        print(f"[WARN] cv2.imshow() unavailable — likely a headless "
                              f"environment/build without GUI support ({e}). "
                              "Continuing without a live preview; the processed "
                              "video will still be written to disk if the writer "
                              "opened successfully.")

                frame1 = frame2
                ret, frame2 = cap.read()

            # Non-blocking key poll (1ms) so the loop / preview never freezes.
            # ESC (27)  AND Q requests a safe stop.
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q') or key == 27:
                break
    

        end_time = time.time()
        execution_time = end_time - start_time
        print("Execution Time: {:.2f} seconds".format(execution_time))

        memory_usage = psutil.Process().memory_info().rss
        print("Memory Usage: {:.2f} MB".format(memory_usage / (1024 * 1024)))

        for lane in (lane1, lane2):
            avg_density = float(np.mean(lane.density_values)) if lane.density_values else 0.0
            print(f"\n-- {lane.name} Summary --")
            print(f"Average Density: {avg_density:.4f}")
            print(f"Vehicles Passed: {lane.vehicle_count}")
            label_counts = Counter(
                rec["label"] for rec in lane.vehicle_history if rec["label"] in ("Car", "Van", "Bus")
            )
            for cls in ("Car", "Van", "Bus"):
                print(f"  {cls}: {label_counts.get(cls, 0)}")

        print("frames: " f"{frame_count}")
    finally:
        # Always released, even if an exception interrupted the loop above,
        # so the capture/writer/window handles never leak.
        cap.release()
        out.release()
        cv2.destroyAllWindows()
        _CELL_POOL.shutdown(wait=True)
        _LANE_POOL.shutdown(wait=True)


def evaluate_lane(lane, gt_data, show_grid_heatmap=True, debug_grid_metrics=False):
    """DETRAC ground-truth evaluation for a single lane — grid-level and
    vehicle-level confusion matrices, precision/recall/F1, and heatmaps.
    Mirrors the evaluation block in seq.py, scoped to this lane's ROI/grid."""
    print(f"\n===== DETRAC Evaluation: {lane.name} =====")

    # Vehicle counts must be ID-based (grid-independent).
    unique_gt_ids = set()
    for _frame_id, objs in gt_data.items():
        roi_gt_boxes = filter_gt_boxes_to_roi(objs, lane.roi_x, lane.roi_y, lane.roi_width, lane.roi_height)
        for obj in roi_gt_boxes:
            obj_id = obj.get("id", -1)
            if obj_id >= 0:
                unique_gt_ids.add(obj_id)

    unique_pred_track_ids = set()
    for _frame_id, pred_vehicles in lane.all_frames_vehicle_data:
        unique_pred_track_ids.update(pred_vehicles.keys())

    print("\n-- Vehicle Count Summary (Grid-Independent) --")
    print(f"Actual vehicles (unique GT IDs): {len(unique_gt_ids)}")
    print(f"Predicted vehicles (unique tracker IDs): {len(unique_pred_track_ids)}")

    gt_vehicle_labels = {}
    for frame_id, objs in gt_data.items():
        roi_gt_boxes = filter_gt_boxes_to_roi(objs, lane.roi_x, lane.roi_y, lane.roi_width, lane.roi_height)
        for obj in roi_gt_boxes:
            vid = obj.get("id", -1)
            label = obj.get("label", "Car")
            if label not in ["Car", "Van", "Bus"]:
                continue
            gt_vehicle_labels[vid] = label

    conf_matrix = None
    if debug_grid_metrics or show_grid_heatmap:
        classes = ["Empty", "Car", "Van", "Bus"]
        class_to_idx = {c: i for i, c in enumerate(classes)}
        conf_matrix = np.zeros((4, 4), dtype=int)

        for frame_id, _binary_grid, pred_grid in lane.all_frames_data:
            gt_boxes = gt_data.get(frame_id + 1, [])
            gt_boxes = filter_gt_boxes_to_roi(gt_boxes, lane.roi_x, lane.roi_y, lane.roi_width, lane.roi_height)
            gt_grid = create_gt_label_grid(
                gt_boxes, lane.roi_x, lane.roi_y, lane.grid_width, lane.grid_height,
                lane.num_rows, lane.num_cols
            )

            for i in range(lane.num_rows):
                for j in range(lane.num_cols):
                    gt_label = gt_grid[i][j]
                    pred_label = pred_grid[i][j]

                    gt_idx = class_to_idx.get(gt_label, 0)
                    pred_idx = class_to_idx.get(pred_label, 0)

                    conf_matrix[gt_idx][pred_idx] += 1

    # Vehicle-level evaluation (bbox vs XML only; not grid-based)
    vehicle_classes = ["Car", "Van", "Bus"]
    vehicle_idx = {c: i for i, c in enumerate(vehicle_classes)}
    conf_matrix_vehicle = np.zeros((3, 3), dtype=int)
    pred_vehicle_labels = {}
    matched_pred_track_ids = set()

    for frame_id, pred_vehicles in lane.all_frames_vehicle_data:
        gt_boxes = gt_data.get(frame_id + 1, [])
        gt_boxes = filter_gt_boxes_to_roi(gt_boxes, lane.roi_x, lane.roi_y, lane.roi_width, lane.roi_height)

        for vid, pred_obj in pred_vehicles.items():
            pred_label = pred_obj.get("label")
            if pred_label is None or pred_label not in vehicle_idx:
                continue

            gt_id = match_vehicle(pred_obj["bbox"], gt_boxes)
            if gt_id is None:
                continue

            pred_vehicle_labels.setdefault(gt_id, []).append(pred_label)
            matched_pred_track_ids.add(vid)

    for gt_id, pred_list in pred_vehicle_labels.items():
        pred_label = Counter(pred_list).most_common(1)[0][0]
        gt_label = gt_vehicle_labels.get(gt_id, "Car")
        if gt_label not in vehicle_idx:
            gt_label = "Car"

        i = vehicle_idx[gt_label]
        j = vehicle_idx[pred_label]
        conf_matrix_vehicle[i][j] += 1

    matched_gt_ids = set(pred_vehicle_labels.keys())
    unmatched_gt_ids = unique_gt_ids - matched_gt_ids
    unmatched_pred_track_ids = unique_pred_track_ids - matched_pred_track_ids

    print("\n-- Vehicle Matching Summary (Unique IDs) --")
    print(f"Matched GT vehicles: {len(matched_gt_ids)}")
    print(f"Unmatched GT vehicles: {len(unmatched_gt_ids)}")
    print(f"Matched predicted tracks: {len(matched_pred_track_ids)}")
    print(f"Unmatched predicted tracks: {len(unmatched_pred_track_ids)}")

    print("\n-- DEBUG INFO --")
    print(f"Total GT vehicles: {len(unique_gt_ids)}")
    print(f"Total predicted vehicles: {len(unique_pred_track_ids)}")
    print(f"Matched GT vehicles: {len(matched_gt_ids)}")
    print(f"Unmatched GT vehicles: {len(unmatched_gt_ids)}")
    print(f"Unmatched predicted vehicles: {len(unmatched_pred_track_ids)}")
    if debug_grid_metrics and conf_matrix is not None:
        print("\n-- DEBUG: Grid Confusion Matrix (Empty-grid heatmap) --")
        print(conf_matrix)

    print("\n-- Vehicle-Level Confusion Matrix --")
    print(conf_matrix_vehicle)

    metrics = compute_vehicle_metrics(conf_matrix_vehicle)
    print("\n-- Vehicle-Level Metrics --")
    for cls in ["Car", "Van", "Bus"]:
        precision, recall, f1 = metrics[cls]
        print(f"\nClass: {cls}")
        print(f"Precision: {precision:.3f}")
        print(f"Recall   : {recall:.3f}")
        print(f"F1 Score : {f1:.3f}")

    if show_grid_heatmap and conf_matrix is not None:
        classes = ["Empty", "Car", "Van", "Bus"]
        plt.figure(figsize=(8, 6))
        sns.heatmap(
            conf_matrix, annot=True, fmt="d", cmap="Blues",
            xticklabels=classes, yticklabels=classes
        )
        plt.title(f"{lane.name} — Multi-Class Grid Confusion Matrix")
        plt.xlabel("Predicted")
        plt.ylabel("Actual")

    classes_vehicle = ["Car", "Van", "Bus"]
    fig, ax = plt.subplots(figsize=(5, 4))
    sns.heatmap(
        conf_matrix_vehicle, annot=True, fmt="d", cmap="viridis",
        xticklabels=classes_vehicle, yticklabels=classes_vehicle, ax=ax,
    )
    ax.set_title(f"{lane.name} — Vehicle-Level Confusion Matrix")
    ax.set_xlabel("Predicted")
    ax.set_ylabel("Actual")

    fig.subplots_adjust(bottom=0.28)
    metrics_text = "\n".join([
        f"{cls}: P={metrics[cls][0]:.3f}  R={metrics[cls][1]:.3f}  F1={metrics[cls][2]:.3f}"
        for cls in classes_vehicle
    ])
    fig.text(0.5, 0.02, metrics_text, ha="center", va="bottom", fontsize=9)

    return conf_matrix, conf_matrix_vehicle, metrics


def save_all_to_excel():
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Lane", "Vehicles Passed", "Average Density"])
    for lane in (lane1, lane2):
        avg_density = float(np.mean(lane.density_values)) if lane.density_values else 0.0
        summary_sheet.append([lane.name, lane.vehicle_count, round(avg_density, 4)])

    # ── Grid Data sheets (binary + label matrix per frame), one per lane ──
    for lane in (lane1, lane2):
        grid_sheet = workbook.create_sheet(f"{lane.name} Grid Data")
        for frame_number, binary_matrix, label_matrix in lane.all_frames_data:
            grid_sheet.append([f"Frame {frame_number} - Binary"])
            for row in binary_matrix:
                grid_sheet.append(list(row))
            grid_sheet.append([])

            grid_sheet.append([f"Frame {frame_number} - Label"])
            for row in label_matrix:
                grid_sheet.append(list(row))
            grid_sheet.append([])

    speed_sheet = workbook.create_sheet("Speed Data")
    speed_sheet.append(["Lane", "Frame", "Vehicle ID", "Label", "Speed (km/h)", "Centroid X", "Centroid Y"])
    for lane in (lane1, lane2):
        for rec in lane.speed_log:
            speed_sheet.append([
                rec["lane"], rec["frame"], rec["vehicle_id"], rec["label"],
                rec["speed_kmph"], rec["centroid_x"], rec["centroid_y"],
            ])

    history_sheet = workbook.create_sheet("Vehicle History")
    history_sheet.append(["Lane", "Vehicle ID", "Final Label", "Avg Speed (km/h)", "Counted"])
    for lane in (lane1, lane2):
        for rec in lane.vehicle_history:
            history_sheet.append([lane.name, rec["vehicle_id"], rec["label"], rec["avg_speed_kmph"], rec["counted"]])

    workbook.save(excel_file_path)


if __name__ == '__main__':
    profiler = cProfile.Profile()
    profiler.enable()
    main()
    profiler.disable()

    stats = pstats.Stats(profiler).sort_stats('cumtime')
    stats.print_stats(10)  # Print the top 10 functions by cumulative time

    # Evaluation runs only after processing has completed.
    XML_PATH = xml_path.strip()
    if XML_PATH and not os.path.isabs(XML_PATH):
        XML_PATH = os.path.abspath(XML_PATH)

    if XML_PATH and os.path.exists(XML_PATH):
        gt_data = load_detrac(XML_PATH)
        for lane in (lane1, lane2):
            evaluate_lane(lane, gt_data, show_grid_heatmap=True, debug_grid_metrics=False)
        os.makedirs("plots", exist_ok=True)
        plt.savefig("plots/evaluation_heatmaps.png")
        plt.close("all")
    elif XML_PATH:
        print(f"Skipping evaluation: XML file not found at {XML_PATH}")
    else:
        print("Skipping evaluation: XML path not provided in user_input_data.json")

    save_all_to_excel()